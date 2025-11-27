from app.repo.report_repo import ReportRepo
from app.repo.permission_repo import PermissionRepo

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import io


class ReportUC:
    def __init__(
        self,
        report_repo=None,
        permission_repo=None,
        intern_repo=None,
        project_repo=None,
        feedback_repo=None
    ):
        # Nếu có report_repo riêng thì dùng, không thì tạo mới
        self.report_repo = report_repo or ReportRepo()
        self.permission_repo = permission_repo or PermissionRepo()
        
        # Backward compatibility với code cũ
        self.intern_repo = intern_repo
        self.project_repo = project_repo
        self.feedback_repo = feedback_repo

    def _check(self, user_id):
        if not self.permission_repo.user_has(user_id, "REPORT_VIEW"):
            raise PermissionError("Missing permission: REPORT_VIEW")

    def export_report(self, user_id, payload):
        self._check(user_id)

        rtype = payload.get("type")
        fdate = payload.get("from_date")
        tdate = payload.get("to_date")
        major = payload.get("major")
        status = payload.get("status")

        filters = {
            "date_from": fdate,
            "date_to": tdate,
            "major": major,
            "status": status
        }

        if rtype == "intern":
            data = self.report_repo.get_interns(filters)
            return self._excel_single(data, "Báo Cáo Sinh Viên Thực Tập")

        if rtype == "project":
            data = self.report_repo.get_projects(filters)
            return self._excel_single(data, "Báo Cáo Dự Án")

        if rtype == "feedback":
            filters["intern_id"] = payload.get("intern_id")
            filters["project_id"] = payload.get("project_id")
            data = self.report_repo.get_feedback(filters)
            return self._excel_single(data, "Báo Cáo Đánh Giá")

        if rtype == "all":
            interns = self.report_repo.get_interns(filters)
            projects = self.report_repo.get_projects(filters)
            feedbacks = self.report_repo.get_feedback(filters)
            return self._excel_multi(interns, projects, feedbacks)

        raise ValueError("Invalid report type")

    def get_statistics(self, user_id, params):
        self._check(user_id)

        filters = {
            "date_from": params.get("from_date"),
            "date_to": params.get("to_date"),
            "major": params.get("major"),
            "status": params.get("status")
        }

        interns = self.report_repo.get_interns(filters)
        projects = self.report_repo.get_projects(filters)
        feedbacks = self.report_repo.get_feedback(filters)

        avg = sum(f.score for f in feedbacks) / len(feedbacks) if feedbacks else 0
        completed = len([p for p in projects if p.status == "done"])
        total = len(projects)
        in_progress = total - completed

        # Đếm theo major
        major_stats = {}
        for i in interns:
            major = i.major or "Chưa phân loại"
            major_stats[major] = major_stats.get(major, 0) + 1

        # Đếm theo trạng thái feedback
        # Phân biệt feedback cho intern (scale 10) và project (scale 5)
        intern_ratings = {"excellent": 0, "good": 0, "average": 0, "poor": 0}
        project_ratings = {"5_star": 0, "4_star": 0, "3_star": 0, "2_star": 0, "1_star": 0}
        
        for f in feedbacks:
            try:
                score = float(f.score) if hasattr(f, 'score') else 0
                
                # Kiểm tra feedback cho intern hay project
                if hasattr(f, 'to_intern_id') and f.to_intern_id:
                    # Feedback cho intern - scale 0-10
                    if score >= 9:
                        intern_ratings["excellent"] += 1
                    elif score >= 7:
                        intern_ratings["good"] += 1
                    elif score >= 5:
                        intern_ratings["average"] += 1
                    else:
                        intern_ratings["poor"] += 1
                        
                elif hasattr(f, 'to_project_id') and f.to_project_id:
                    # Feedback cho project - scale 1-5 sao
                    if score >= 4.5:
                        project_ratings["5_star"] += 1
                    elif score >= 3.5:
                        project_ratings["4_star"] += 1
                    elif score >= 2.5:
                        project_ratings["3_star"] += 1
                    elif score >= 1.5:
                        project_ratings["2_star"] += 1
                    else:
                        project_ratings["1_star"] += 1
                        
            except (ValueError, TypeError, AttributeError):
                intern_ratings["poor"] += 1

        return {
            "intern_count": len(interns),
            "project_total": total,
            "project_completed": completed,
            "project_in_progress": in_progress,
            "project_completion_rate": completed / total if total else 0,
            "average_rating": round(avg, 2) if avg else 0,
            "feedback_count": len(feedbacks),
            "major_stats": major_stats,
            "intern_rating_distribution": intern_ratings,
            "project_rating_distribution": project_ratings
        }

    def view_report(self, user_id, filters):
        self._check(user_id)

        rtype = filters.get("type")
        fdate = filters.get("from_date")
        tdate = filters.get("to_date")
        major = filters.get("major")
        status = filters.get("status")

        query_filters = {
            "date_from": fdate,
            "date_to": tdate,
            "major": major,
            "status": status
        }

        if rtype == "intern":
            return self.report_repo.get_interns(query_filters)
        if rtype == "project":
            return self.report_repo.get_projects(query_filters)
        if rtype == "feedback":
            query_filters["intern_id"] = filters.get("intern_id")
            query_filters["project_id"] = filters.get("project_id")
            return self.report_repo.get_feedback(query_filters)

        raise ValueError("Invalid type")

    def _excel_single(self, records, title):
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit

        # Header
        ws.merge_cells("A1:E1")
        ws["A1"] = title
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="4472C4", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        if not records:
            ws.append(["Không có dữ liệu"])
        else:
            data_dicts = [self._clean_dict(r.to_dict()) for r in records]
            keys = list(data_dicts[0].keys())

            # Column headers
            ws.append(keys)
            for c in ws[2]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color="5B9BD5", fill_type="solid")
                c.alignment = Alignment(horizontal="center", vertical="center")

            # Data rows
            for rec in data_dicts:
                ws.append([rec.get(k) for k in keys])

        self._format_sheet(ws)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def _excel_multi(self, interns, projects, feedbacks):
        wb = Workbook()

        sheets = {
            "Sinh viên": interns,
            "Dự án": projects,
            "Đánh giá": feedbacks
        }

        first = True
        for sheet_name, items in sheets.items():
            ws = wb.active if first else wb.create_sheet(sheet_name)
            if not first:
                ws.title = sheet_name
            first = False

            # Header
            ws.merge_cells("A1:E1")
            ws["A1"] = f"Báo Cáo {sheet_name}"
            ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
            ws["A1"].fill = PatternFill(start_color="4472C4", fill_type="solid")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25

            if not items:
                ws.append(["Không có dữ liệu"])
            else:
                data_dicts = [self._clean_dict(i.to_dict()) for i in items]
                keys = list(data_dicts[0].keys())

                ws.append(keys)
                for c in ws[2]:
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = PatternFill(start_color="5B9BD5", fill_type="solid")
                    c.alignment = Alignment(horizontal="center", vertical="center")

                for row in data_dicts:
                    ws.append([row.get(k) for k in keys])

            self._format_sheet(ws)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def _format_sheet(self, ws):
        thin = Side(border_style="thin", color="CCCCCC")

        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col in ws.columns:
            length = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 3, 50)

    def _clean_dict(self, d):
        return {k: v for k, v in d.items() if k not in ["is_deleted", "password_hash"]}